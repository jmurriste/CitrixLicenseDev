
import pprint
from openpyxl import Workbook


import json
from bson import json_util

from pymongo import MongoClient
client = MongoClient('7.188.0.80', 37777)
db = client['Citrix']
collection = db['Costs']
coll = db.list_collection_names()
print(len(coll))

a = []
val =[]
for i in range(0, len(coll)):
        val = [i,coll[i]]
        a.append(val)
print(a)


CostsCursor = collection.find({})
wb = Workbook()
sheet1 = wb.active



print(a)

row = 2

sheet1.cell(row=1, column=1).value = "License Product Name"
sheet1.cell(row=1, column=2).value = "License SKU"
sheet1.cell(row=1, column=3).value = "License Price"
sheet1.cell(row=1, column=4).value = "Type"
for i in CostsCursor:
    print(i['License Product Name'])
    sheet1.cell(row=row, column=1).value = i['License Product Name']
    sheet1.cell(row=row, column=2).value = i['License SKU']
    sheet1.cell(row=row, column=3).value = i['License Price']
    sheet1.cell(row=row, column=4).value = i['Type']
    row+=1

wb.save(filename='valores.xlsx')

json_docs = [json.dumps(doc, default=json_util.default) for doc in CostsCursor]

#pprint.pprint(json_docs)



wb = Workbook()
sheet1 = wb.active
#sheet1.cell(row=k, column=ncolumna).value = i[0]





