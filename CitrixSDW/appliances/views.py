from flask import Blueprint,render_template, request, send_file
from CitrixSDW import mongo
from CitrixSDW.appliances.forms import DownloadForm, ReporteForm
from openpyxl import Workbook


appliances_blueprint = Blueprint('appliances',
                              __name__,
                              template_folder='templates/appliances')


@appliances_blueprint.route('/download', methods=['GET', 'POST'])
def download():
    citrix = mongo.db
    dbnames = citrix.list_collection_names({})
    choices = []
    for i in range(0, len(dbnames)):
        val = [dbnames[i], dbnames[i]]
        choices.append(val)

    form = DownloadForm()
    form.reportes.choices = choices

    if request.method == 'POST':

        dbselected = form.reportes.data

        if dbselected == 'Costs':

            citrix = mongo.db.Costs
            CostsCursor = citrix.find({})

            wb = Workbook()
            sheet1 = wb.active
            row = 2

            sheet1.cell(row=1, column=1).value = "License Product Name"
            sheet1.cell(row=1, column=2).value = "License SKU"
            sheet1.cell(row=1, column=3).value = "License Price"
            sheet1.cell(row=1, column=4).value = "Type"

            for i in CostsCursor:

                sheet1.cell(row=row, column=1).value = i['License Product Name']
                sheet1.cell(row=row, column=2).value = i['License SKU']
                sheet1.cell(row=row, column=3).value = i['License Price']
                sheet1.cell(row=row, column=4).value = i['Type']
                row += 1

            wb.save(filename=str(dbselected)+'.xlsx')
            return send_file('C:\\Users\\jorge\\Dropbox (Neutrona)\\JORGE\\PYTHON CODE\\Flask\\Netscaler\\' + str(
                dbselected) + '.xlsx',
                             mimetype='text/csv',
                             attachment_filename=str(dbselected) + '.xlsx',
                             as_attachment=True)


        else:
            from pymongo import MongoClient
            client = MongoClient('7.188.0.80', 37777)
            db = client['Citrix']
            collection = db[dbselected]
            ReportCursor = collection.find()

            wb = Workbook()
            sheet2 = wb.active
            row = 2

            sheet2.cell(row=1, column=1).value = "name"
            sheet2.cell(row=1, column=2).value = "license_state"
            sheet2.cell(row=1, column=3).value = "model"
            sheet2.cell(row=1, column=4).value = "serial_no"
            sheet2.cell(row=1, column=5).value = "product"


            sheet2.cell(row=1, column=6).value = "license_type"
            sheet2.cell(row=1, column=7).value = "local_license_server_hostid"
            sheet2.cell(row=1, column=8).value = "license_expiry"
            sheet2.cell(row=1, column=9).value = "max_bw"
            sheet2.cell(row=1, column=10).value = "model"
            sheet2.cell(row=1, column=11).value = "state"
            sheet2.cell(row=1, column=12).value = "system_patform"

            sheet2.cell(row=1, column=13).value = "description"
            sheet2.cell(row=1, column=14).value = "country"
            sheet2.cell(row=1, column=15).value = "city"
            sheet2.cell(row=1, column=16).value = "POP"
            sheet2.cell(row=1, column=17).value = "SID"
            sheet2.cell(row=1, column=18).value = "BW Report (Mbps)"
            sheet2.cell(row=1, column=19).value = "Error"


            for i in ReportCursor:

                try:
                    sheet2.cell(row=row, column=1).value = i['name']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=2).value = i['license_state']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=3).value = i['model']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=4).value = i['serial_no']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=5).value = i['product']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=6).value = i['license_info']['license_type']

                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=7).value = i['license_info']['local_license_server_hostid']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=8).value = i['license_info']['license_expiry']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=9).value = i['license_info']['max_bw']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=10).value = i['license_info']['model']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=11).value = i['license_info']['state']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=12).value = i['license_info']['system_patform']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=13).value = i['description']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=14).value = i['country']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=15).value = i['city']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=16).value = i['POP']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=17).value = i['SID']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=18).value = i['BW Report (Mbps)']
                except Exception as e:
                    pass
                try:
                    sheet2.cell(row=row, column=19).value = i['error']
                except Exception as e:
                    pass

                row += 1
            wb.save(filename=str(dbselected) + '.xlsx')
            return send_file('C:\\Users\\jorge\\Dropbox (Neutrona)\\JORGE\\PYTHON CODE\\Flask\\Netscaler\\'+str(dbselected) + '.xlsx',
                             mimetype='text/csv',
                             attachment_filename=str(dbselected) + '.xlsx',
                             as_attachment=True)

        return '<h1> db: {}</h1>'.format(dbselected)

    return render_template('download.html', form=form)

@appliances_blueprint.route('/costs')
def costs():

    citrix = mongo.db.Costs
    CostsCursor = citrix.find({})

    return render_template('costs.html', CostsCursor=CostsCursor)


@appliances_blueprint.route('/list', methods=['GET', 'POST'])
def list():

    citrix = mongo.db
    dbnames = citrix.list_collection_names({})
    choices = []
    for i in range(0, len(dbnames)):
        if dbnames[i] == 'Costs':
            pass
        else:
            val = [dbnames[i], dbnames[i]]
            choices.append(val)

    form = ReporteForm()
    form.reportes.choices = choices

    if request.method == 'POST':
        dbselected = form.reportes.data

        from pymongo import MongoClient
        client = MongoClient('7.188.0.80', 37777)
        db = client['Citrix']
        collection = db[dbselected]
        ReportCursor = collection.find()

        newcursor = []

        for i in ReportCursor:

            try:
                license_type = {'license_type': i['license_info']['license_type']}
                i.update(license_type)
            except Exception as e:
                pass
            try:
                local_license_server_hostid = {
                    'local_license_server_hostid': i['license_info']['local_license_server_hostid']}
                i.update(local_license_server_hostid)
            except Exception as e:
                pass
            try:
                license_expiry = {'license_expiry': i['license_info']['license_expiry']}
                i.update(license_expiry)
            except Exception as e:
                pass
            try:
                max_bw = {'max_bw': i['license_info']['max_bw']}
                i.update(max_bw)
            except Exception as e:
                pass
            try:
                model = {'model': i['license_info']['model']}
                i.update(model)
            except Exception as e:
                pass
            try:
                state = {'state': i['license_info']['state']}
                i.update(state)
            except Exception as e:
                pass
            try:
                system_patform = {'system_patform': i['license_info']['system_patform']}
                i.update(system_patform)
            except Exception as e:
                pass
            i.pop('license_info', None)
            newcursor.append(i)
            values = len(newcursor)

        return render_template('list.html', form=form, values=values, newcursor=newcursor)

    return render_template('list-simple.html', form=form)

